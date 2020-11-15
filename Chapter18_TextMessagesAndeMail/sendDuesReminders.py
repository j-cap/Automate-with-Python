
# sendDuesReminders.py - Sends emails based on payment status in spreadsheets.

import openpyxl, smtplib, sys
# Open the spreadsheet and get the lastest dues status
wb = openpyxl.load_workbook("../materials/duesRecords.xlsx")
sheet = wb["Sheet1"]
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# TODO: Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != "paid":
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
# TODO: Login to email account
smtpObj = smtplib.SMTP("smtp.gmail.com", 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login("weberjakob64@gmail.com", sys.argv[1])

# TODO: Send out reminder emails.
for name, email in unpaidMembers.items():
    body = f"Subject: {latestMonth} dues unpaid. \nDear {name}, \n Records show that you have not paid dues for {latestMonth}. Please make this payment as soon as possible. Thank you!"
    print("Sending email to %s..."%email)
    sendmailStatus = smtpObj.sendmail("weberjakob64@gmail.com", email, body)
    if sendmailStatus != {}:
        print(f"There was a problem sending mail 